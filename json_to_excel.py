import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

JSON_PATH = "output/comparison.json"
EXCEL_PATH = "output/comparison.xlsx"

# Load JSON
with open(JSON_PATH, "r", encoding="utf-8") as f:
    data = json.load(f)

rows = data["comparison"]

# Create DataFrame
df = pd.DataFrame(rows)

# Convert TRUE/FALSE → Yes/No
df = df.replace({True: "Yes", False: "No"})

# Ensure output folder exists
os.makedirs("output", exist_ok=True)

# Write Excel
df.to_excel(EXCEL_PATH, index=False)

# -------------------------------
# CONDITIONAL FORMATTING
# -------------------------------

wb = load_workbook(EXCEL_PATH)
ws = wb.active

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Apply formatting (skip first column: pointer_theme)
for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column):
    for cell in row:
        if cell.value == "Yes":
            cell.fill = green_fill
        elif cell.value == "No":
            cell.fill = red_fill

wb.save(EXCEL_PATH)

print("✅ Excel created with green/red conditional formatting")
